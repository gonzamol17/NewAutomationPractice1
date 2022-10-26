import time
import pytest
import unittest
import sys
import os

from Data.DataLoginUser import DataLoginUser
from POM.RegistrationPage import RegistrationPage

sys.path.append(os.path.join(os.path.dirname(__file__),"..",".."))
import json
import openpyxl
from colorama import Fore, Back, Style
import HtmlTestRunner
from Utils import utils as utils
from POM.HomePage import HomePage
from POM.LoginPage import LoginPage
from Utils.BaseClass import BaseClass

@pytest.mark.usefixtures("test_setup")
class TestReadInformationFromExcel(BaseClass):

    def test_Verify_Read_Information_From_Excel(self, getData2):
        # book = openpyxl.load_workbook("C:\\Users\\admin\\PycharmProjects\\NewAutomationPractice\\Data\\Libro1.xlsx")
        # sheet = book.active
        # # cell = sheet.cell(row=1, column=1)
        # # print(cell.value)
        # # print(sheet.max_row)
        # # print(sheet.max_column)
        # for i in range(1, sheet.max_row+1):
        #     for j in range(2, sheet.max_column+1):
        #         print(sheet.cell(row=i, column=j).value)
        log = self.get_Logger()
        driver = self.driver
        hp = HomePage(driver)
        time.sleep(3)
        log.info("Se hace click en link signIn")
        hp.select_SignIn()
        time.sleep(3)
        lp = LoginPage(driver)
        log.info("Se ingresa email y password")
        lp.do_Login(getData2['Email'], getData2['Password'])
        assert "Logged in as" in hp.verify_Logged()
        log.info("El usuario está logueado")
        time.sleep(3)




    @pytest.fixture(params=DataLoginUser.getTestData("User1"))
    def getData2(self, request):
        return request.param


