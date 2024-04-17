import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


class DataLoginUser:

        @staticmethod
        def getTestData(test_case_name):
            book = openpyxl.load_workbook("C:\\Users\\User\\PycharmProjects\\NewAutomationPractice1\\Data\\Libro1.xlsx")
            sheet = book.active
            Dict= {}
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == test_case_name:
                    for j in range(2, sheet.max_column + 1):
                        #print(sheet.cell(row=i, column=j).value)
                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return [Dict]
